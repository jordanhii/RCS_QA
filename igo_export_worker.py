#!/usr/bin/env python3
"""
igo_export_worker.py — IGO Specialty Games 批量导出工具（两阶段并发版）

Phase 1: Playwright 打开浏览器，拦截汇总 API 请求（URL / Headers / Body）
Phase 2: aiohttp 并发复放所有历史日期的请求（替换日期字符串）

导出内容：单个 Excel 文件，包含以下 sheet（根据参数选择）:
  - 原始数据 (exportRaw=true)
  - 每日数据 (exportAnalysis=true)
  - 告警分析 (exportAnalysis=true)

用法：
  python3 igo_export_worker.py \\
    --query-date  2026-05-28 \\
    --end-time    11:00:59 \\
    --days-back   30 \\
    --outlet-ids  "IGO,Lavie,NCLI,Stotsenberg" \\
    --game-brands "IGO-COLORGAME,QAT-55-COLORGAME,QAT-COLORGAME" \\
    --x-threshold 5.0 \\
    --y-threshold 10.0 \\
    --prev-ggr    0 \\
    --export-raw \\
    --export-analysis \\
    --output /tmp/igo_export.xlsx
"""

import argparse
import asyncio
import json
import os
import re
import statistics
import sys
from datetime import datetime, timedelta, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

try:
    import aiohttp
    HAS_AIOHTTP = True
except ImportError:
    HAS_AIOHTTP = False

# ─── 常量 ─────────────────────────────────────────────────────────────────────
IGO_BASE_URL = "https://igo-web.igo8.me/igo-report"
TARGET_PAGE  = f"{IGO_BASE_URL}/transaction/specialtyGames"
STATE_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "igo_state.json")

# 账号不再写死：优先用后端注入的 RCS_CRED_*（数据库账号配置，已解密），
# 否则用环境变量 IGO_USERNAME/IGO_PASSWORD；仍可被 --username/--password 覆盖。
DEFAULT_USERNAME = os.environ.get("RCS_CRED_USER") or os.environ.get("IGO_USERNAME", "")
DEFAULT_PASSWORD = os.environ.get("RCS_CRED_PASS") or os.environ.get("IGO_PASSWORD", "")

NUMERIC_FIELDS = [
    "totalAbandonedPot", "totalAmount", "totalBetCash", "totalBetNN",
    "totalBonusAmount", "totalCount", "totalHouseRake",
    "totalJackpotContribution", "totalJackpotContributionSup",
    "totalJackpotPayout", "totalNonAvailableAmount",
    "totalPayout", "totalPotFee", "totalValidAmount",
]
SUMMARY_KEYS = {"totalAmount", "totalCount", "totalBetCash", "totalValidAmount"}
MAX_CONCURRENT = 8

# ─── 日志 & 错误 ──────────────────────────────────────────────────────────────
def plog(msg: str) -> None:
    print(f"[IGO导出] {msg}", flush=True)


def err_exit(msg: str, hint: str = "") -> None:
    payload = {"error": msg}
    if hint:
        payload["hint"] = hint
    print(json.dumps(payload, ensure_ascii=False), file=sys.stderr)
    sys.exit(1)


# ─── 响应识别 ─────────────────────────────────────────────────────────────────
def is_summary(obj: dict) -> bool:
    inner = obj.get("data", obj)
    if not isinstance(inner, dict):
        return False
    has_sum = bool(SUMMARY_KEYS & set(inner.keys()))
    has_rec = ("records" in inner
               and isinstance(inner.get("records"), list)
               and len(inner.get("records", [])) > 0)
    return has_sum and not has_rec


# ─── 请求复放辅助 ─────────────────────────────────────────────────────────────
def build_request_for_day(captured: dict, day: date, orig_day: date) -> dict:
    old_date = str(orig_day)
    new_date = str(day)
    url  = captured["url"].replace(old_date, new_date)
    data = captured.get("data", b"")
    if data:
        data = data.decode("utf-8", errors="replace").replace(old_date, new_date).encode("utf-8")
    return {"url": url, "method": captured["method"], "headers": captured["headers"], "data": data}


# ─── 登录处理 ─────────────────────────────────────────────────────────────────
async def handle_login(page, username: str, password: str) -> bool:
    await asyncio.sleep(1)
    if await page.locator('input[type="password"]').count() == 0:
        plog("✅ 会话有效，已登录")
        return True
    plog("检测到登录页，正在自动登录...")
    try:
        user_sel = ('input[type="text"],input[type="email"],'
                    'input[placeholder*="user" i],input[placeholder*="account" i]')
        await page.locator(user_sel).first.fill(username, timeout=5000)
        await page.locator('input[type="password"]').first.fill(password, timeout=5000)
        btn_sel = ('button[type="submit"],'
                   'button:text-matches("login|sign in|登录|确定","i"),'
                   'input[type="submit"]')
        await page.locator(btn_sel).first.click(timeout=5000)
        await page.wait_for_function(
            "() => !location.href.toLowerCase().includes('login')", timeout=20000)
        plog("✅ 登录成功")
        return True
    except Exception as e:
        plog(f"❌ 登录失败: {e}")
        return False


# ─── 多选字段设置 ─────────────────────────────────────────────────────────────
async def set_multiselect(page, nth: int, label: str, values: list) -> None:
    """设置 IGO 的 custom-select 多选字段（Outlet ID / Game Brand）"""
    if not values:
        return
    plog(f"设置 [{label}]: {values}")
    inp_loc = page.locator("input.custom-select__input")
    if await inp_loc.count() <= nth:
        plog(f"⚠ 找不到 [{label}] 输入框，跳过")
        return

    # 标记目标容器
    await page.evaluate(f"""
        (nth) => {{
            const inputs = document.querySelectorAll('input.custom-select__input');
            if (!inputs[nth]) return;
            let el = inputs[nth];
            for (let i = 0; i < 6; i++) {{
                el = el.parentElement; if (!el) return;
                if ((el.className || '').includes('custom-select') &&
                    !(el.className || '').includes('custom-select__')) {{
                    el.setAttribute('data-cst', nth); return;
                }}
            }}
            inputs[nth].parentElement.setAttribute('data-cst', nth);
        }}
    """, nth)
    cont = page.locator(f'[data-cst="{nth}"]')

    # 清除已选值
    for _ in range(20):
        btns = cont.locator('[class*="close"],[class*="remove"],[class*="delete"]')
        if await btns.count() == 0:
            break
        try:
            await btns.first.click(timeout=800)
            await asyncio.sleep(0.2)
        except Exception:
            break

    inp = inp_loc.nth(nth)
    for val in values:
        await inp.click(timeout=3000)
        await asyncio.sleep(0.5)
        # 清空输入框
        await page.evaluate("""
            (nth) => {
                const inputs = document.querySelectorAll('input.custom-select__input');
                if (!inputs[nth]) return;
                Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')
                      .set.call(inputs[nth], '');
            }
        """, nth)
        await asyncio.sleep(0.2)
        await page.keyboard.type(val, delay=80)
        await asyncio.sleep(1.5)
        # 点击匹配的下拉选项
        result = await page.evaluate("""
            (s) => {
                const UP = t => t.toUpperCase();
                const cands = Array.from(document.querySelectorAll('*')).filter(el => {
                    if (el.children.length > 3) return false;
                    const t = el.textContent.trim(); if (!t) return false;
                    const ok = (t === s || UP(t) === UP(s)) ||
                               (t.toLowerCase().startsWith(s.toLowerCase()) && t.length - s.length <= 8);
                    if (!ok) return false;
                    const r = el.getBoundingClientRect();
                    return r.width > 0 && r.height > 0;
                });
                if (!cands.length) return 'not_found';
                cands.sort((a, b) => a.children.length - b.children.length);
                const ex = cands.find(el => el.textContent.trim() === s ||
                                            el.textContent.trim().toUpperCase() === s.toUpperCase());
                const t = ex || cands[0]; t.click();
                return 'clicked:' + t.textContent.trim();
            }
        """, val)
        plog(f"  [{val}] {result}")
        await asyncio.sleep(0.6 if str(result).startswith("clicked:") else 0.3)

    await page.keyboard.press("Escape")
    await asyncio.sleep(0.3)
    await page.evaluate(f"document.querySelectorAll('[data-cst=\"{nth}\"]')"
                        f".forEach(el => el.removeAttribute('data-cst'))")


# ─── 日期时间范围设置 ─────────────────────────────────────────────────────────
async def set_datetime_range(page, day: date, start_time: str, end_time: str) -> bool:
    ss = f"{day} {start_time}"
    es = f"{day} {end_time}"
    try:
        result = await page.evaluate(f"""
            () => {{
                const phs = ['Payout Time Start', 'Payout Time End'];
                let found = [];
                for (const ph of phs) {{
                    const el = document.querySelector('input[placeholder="' + ph + '"]');
                    if (el) found.push(el);
                }}
                if (found.length < 2) {{
                    const all = Array.from(document.querySelectorAll('input'));
                    const dated = all.filter(i => /\\d{{4}}-\\d{{2}}-\\d{{2}}/.test(i.value));
                    if (dated.length >= 2) found = dated;
                    else if (all.length >= 2) found = all;
                }}
                if (found.length < 2) return 'NOT_FOUND';
                const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
                setter.call(found[0], '{ss}');
                found[0].dispatchEvent(new Event('input', {{bubbles: true}}));
                found[0].dispatchEvent(new Event('change', {{bubbles: true}}));
                setter.call(found[1], '{es}');
                found[1].dispatchEvent(new Event('input', {{bubbles: true}}));
                found[1].dispatchEvent(new Event('change', {{bubbles: true}}));
                return 'OK:' + found[0].value;
            }}
        """)
        if result and str(result).startswith("OK:"):
            await asyncio.sleep(0.4)
            await page.keyboard.press("Escape")
            return True
    except Exception as ex:
        plog(f"⚠ 时间设置出错: {ex}")
    return False


# ─── Phase 1: 浏览器抓取首日数据 & API 请求 ──────────────────────────────────
async def phase1_capture(target_date: date, start_time: str, end_time: str,
                         outlet_ids: list, game_brands: list,
                         username: str, password: str) -> dict:
    """使用 Playwright 导航、登录、设置筛选条件、拦截汇总 API 请求"""
    captured_req  = {}
    first_summary = {}

    use_state = os.path.exists(STATE_FILE)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--no-sandbox"])
        context = (await browser.new_context(storage_state=STATE_FILE)
                   if use_state else await browser.new_context())
        page = await context.new_page()
        summary_event = asyncio.Event()

        async def on_response(resp):
            if first_summary:
                return
            if resp.status != 200:
                return
            if "json" not in resp.headers.get("content-type", ""):
                return
            try:
                obj = await resp.json()
            except Exception:
                return
            if is_summary(obj):
                inner = obj.get("data", obj)
                first_summary.update(inner)
                req  = resp.request
                body = b""
                try:
                    body = await req.body()
                except Exception:
                    pass
                captured_req.update({
                    "url":     req.url,
                    "method":  req.method,
                    "headers": dict(req.headers),
                    "data":    body,
                })
                summary_event.set()
                plog(f"📡 截获汇总接口  BET={inner.get('totalAmount', 0):,.0f}")
                plog(f"📌 {req.method} {req.url[:80]}")

        page.on("response", on_response)

        plog("正在导航到 Specialty Games...")
        try:
            await page.goto(TARGET_PAGE, wait_until="networkidle", timeout=30000)
        except Exception:
            pass

        if not await handle_login(page, username, password):
            await browser.close()
            raise RuntimeError("IGO 登录失败，请检查账号密码")

        if "specialtyGames" not in page.url:
            plog("重新导航到目标页面...")
            try:
                await page.goto(TARGET_PAGE, wait_until="networkidle", timeout=30000)
            except Exception:
                pass
        await asyncio.sleep(2)

        try:
            await context.storage_state(path=STATE_FILE)
            plog("✅ Session 已保存")
        except Exception:
            pass

        plog(f"设置时间范围: {target_date} {start_time} ~ {end_time}")
        await set_datetime_range(page, target_date, start_time, end_time)
        await asyncio.sleep(0.5)
        await set_multiselect(page, 0, "Outlet ID",  outlet_ids)
        await asyncio.sleep(0.5)
        await set_multiselect(page, 1, "Game Brand", game_brands)
        await asyncio.sleep(0.5)

        plog("点击查询...")
        for btn_text in ["Search", "Query", "查询"]:
            try:
                btn = page.get_by_role("button", name=re.compile(btn_text, re.IGNORECASE))
                if await btn.count() > 0:
                    await btn.first.click(timeout=4000)
                    plog(f"✅ 已点击「{btn_text}」")
                    break
            except Exception:
                pass

        try:
            await asyncio.wait_for(summary_event.wait(), timeout=25)
        except asyncio.TimeoutError:
            plog("⚠ 等待汇总响应超时（25s）")

        cookies = await context.cookies()
        plog(f"✅ 已提取 {len(cookies)} 个 Cookie")
        await browser.close()
        plog("浏览器已关闭")

    if not captured_req:
        raise RuntimeError("未能捕获汇总 API 请求，请检查筛选条件或时间范围是否有数据")

    return {
        "captured":   captured_req,
        "cookies":    cookies,
        "first_data": first_summary if first_summary else None,
    }


# ─── Phase 2: aiohttp 并发复放历史请求 ───────────────────────────────────────
async def phase2_parallel(phase1: dict, days: list, target_date: date) -> dict:
    """将已捕获的 API 请求按日期逐一复放，并发度 MAX_CONCURRENT"""
    if not HAS_AIOHTTP:
        plog("⚠ 未安装 aiohttp，跳过 Phase 2（只有查询日期的数据）")
        return {}

    captured  = phase1["captured"]
    cookies   = {c["name"]: c["value"] for c in phase1["cookies"]}
    skip_hdrs = {"cookie", "content-length", "host"}
    headers   = {k: v for k, v in captured["headers"].items()
                 if k.lower() not in skip_hdrs}

    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    results: dict = {}

    async def fetch_one(session, day: date) -> None:
        req = build_request_for_day(captured, day, target_date)
        async with semaphore:
            try:
                method = req["method"].upper()
                kwargs = dict(headers=headers, ssl=False,
                              timeout=aiohttp.ClientTimeout(total=30))
                if method == "POST":
                    kwargs["data"] = req["data"]
                    resp_obj = await session.post(req["url"], **kwargs)
                else:
                    resp_obj = await session.get(req["url"], **kwargs)

                if resp_obj.status == 200:
                    try:
                        obj = await resp_obj.json(content_type=None)
                        if is_summary(obj):
                            results[str(day)] = obj.get("data", obj)
                            plog(f"  ✅ {day}  BET={results[str(day)].get('totalAmount', 0):,.0f}")
                            return
                    except Exception:
                        pass
                plog(f"  ⚠ {day}  HTTP {resp_obj.status}")
                results[str(day)] = None
            except Exception as e:
                plog(f"  ❌ {day}  {e}")
                results[str(day)] = None

    plog(f"Phase 2：并发复放 {len(days)} 天请求（并发={MAX_CONCURRENT}）...")
    async with aiohttp.ClientSession(cookies=cookies) as session:
        await asyncio.gather(*[fetch_one(session, d) for d in days])

    return results


# ─── 告警计算 ─────────────────────────────────────────────────────────────────
def sf(v):
    try:
        return float(v) if v not in ("", None) else 0.0
    except Exception:
        return 0.0


def row_rtp(r: dict) -> float:
    b = sf(r.get("totalAmount", 0))
    p = sf(r.get("totalPayout", 0))
    return ((b - p) / b * 100) if b > 0 else 0.0


def compute_analysis(rows: list, target_date_str: str,
                     x_threshold: float, y_threshold: float,
                     prev_ggr: float, days_back: int) -> dict:
    today_row = next((r for r in rows if r.get("日期") == target_date_str), None)
    past_rows = [r for r in rows if r.get("日期") != target_date_str]

    cur_bet    = sf(today_row.get("totalAmount",    0)) if today_row else 0.0
    cur_payout = sf(today_row.get("totalPayout",    0)) if today_row else 0.0
    ggr_raw    = sf(today_row.get("totalBonusAmount", 0)) if today_row else 0.0
    cur_ggr    = -ggr_raw  # 取反
    cur_rtp    = ((cur_bet - cur_payout) / cur_bet * 100) if cur_bet > 0 else 0.0

    past_bets  = [sf(r.get("totalAmount", 0)) for r in past_rows]
    past_rtps  = [row_rtp(r) for r in past_rows]
    sorted_bets    = sorted(past_bets)
    avg_nd_rtp     = (sum(past_rtps) / days_back) if past_rtps else 0.0
    median_nd_bet  = statistics.median(past_bets) if past_bets else 0.0

    rtp_diff = cur_rtp - avg_nd_rtp
    alert1   = rtp_diff < x_threshold
    alert2   = cur_bet >= median_nd_bet
    c3_diff  = prev_ggr - cur_ggr
    c3_abs   = abs(prev_ggr) * (y_threshold / 100)
    alert3   = c3_diff >= c3_abs

    return dict(
        target_date=target_date_str,
        cur_bet=cur_bet, cur_payout=cur_payout, cur_rtp=cur_rtp,
        cur_ggr=cur_ggr, ggr_raw=ggr_raw,
        sorted_bets=sorted_bets,
        avg_nd_rtp=avg_nd_rtp, median_nd_bet=median_nd_bet,
        rtp_diff=rtp_diff,
        x_threshold=x_threshold, y_threshold=y_threshold,
        prev_ggr=prev_ggr,
        c3_diff=c3_diff, c3_abs=c3_abs,
        alert1=alert1, alert2=alert2, alert3=alert3,
        normal_alert=alert1 and alert2,
        consecutive_alert=alert1 and alert2 and alert3,
        days_back=days_back,
    )


# ─── Excel 样式辅助 ───────────────────────────────────────────────────────────
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right():
    return Alignment(horizontal="right", vertical="center")


# ─── Sheet 1：原始数据 ────────────────────────────────────────────────────────
def write_raw_sheet(ws, rows: list, start_time: str, end_time: str) -> None:
    navy = PatternFill("solid", fgColor="1F3864")
    gold = PatternFill("solid", fgColor="C9A227")
    thin = _thin()
    all_cols = ["日期", "时间段", "状态"] + NUMERIC_FIELDS
    for ci, h in enumerate(all_cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = navy
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = _center()
        c.border = thin
    ws.row_dimensions[1].height = 30
    for ri, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="EFF3FB" if ri % 2 == 0 else "FFFFFF")
        for ci, col in enumerate(all_cols, 1):
            val = row.get(col, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.border = thin
            c.alignment = _center() if ci <= 3 else _right()
            if ci > 3 and isinstance(val, (int, float)):
                c.number_format = "#,##0.000"
    sr = len(rows) + 2
    for ci in range(1, len(all_cols) + 1):
        c = ws.cell(row=sr, column=ci)
        c.fill = gold
        c.font = Font(bold=True)
        c.border = thin
        c.alignment = _right()
    ws.cell(row=sr, column=1, value="合计").alignment = _center()
    ws.cell(row=sr, column=2, value=f"{start_time}~{end_time}").alignment = _center()
    for ci, field in enumerate(NUMERIC_FIELDS, 4):
        vals = [r[field] for r in rows if isinstance(r.get(field), (int, float))]
        c = ws.cell(row=sr, column=ci, value=sum(vals))
        c.number_format = "#,##0.000"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    for ci in range(4, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22


# ─── Sheet 2：每日数据 ────────────────────────────────────────────────────────
def write_daily_sheet(ws, rows: list, ana: dict,
                      start_time: str, end_time: str) -> None:
    navy = PatternFill("solid", fgColor="1F3864")
    yell = PatternFill("solid", fgColor="FFF2CC")
    thin = _thin()
    hdrs = ["日期", "时间段", "BET", "PAYOUT", "RTP (%)", "备注"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = navy
        c.font = Font(color="FFFFFF", bold=True)
        c.border = thin
        c.alignment = _center()
    ws.row_dimensions[1].height = 25
    for ri, row in enumerate(rows, 2):
        d   = row.get("日期", "")
        b   = sf(row.get("totalAmount", 0))
        p_v = sf(row.get("totalPayout", 0))
        rtp = ((b - p_v) / b * 100) if b > 0 else 0.0
        is_t = (d == ana["target_date"])
        fill = yell if is_t else PatternFill("solid", fgColor="EFF3FB" if ri % 2 == 0 else "FFFFFF")
        vals = [d, f"{start_time}~{end_time}", round(b, 3), round(p_v, 3),
                round(rtp, 4), "★ 查询日期" if is_t else ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.border = thin
            c.alignment = _center() if ci in (1, 2, 6) else _right()
            if ci in (3, 4):
                c.number_format = "#,##0.000"
            if ci == 5:
                c.number_format = "#,##0.0000"
    for col, w in zip("ABCDEF", [14, 22, 18, 18, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# ─── Sheet 3：告警分析 ────────────────────────────────────────────────────────
def write_analysis_sheet(ws, ana: dict, start_time: str, end_time: str) -> None:
    navy  = PatternFill("solid", fgColor="1F3864")
    red   = PatternFill("solid", fgColor="FFCCCC")
    green = PatternFill("solid", fgColor="CCFFCC")
    sect  = PatternFill("solid", fgColor="D9E1F2")
    thin  = _thin()
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 32

    def wrow(r, label, value, fill=None, bold=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        for c in (c1, c2):
            c.font = Font(bold=bold)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                c.fill = fill
        ws.row_dimensions[r].height = 20

    def shdr(r, title):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = title
        c.font  = Font(bold=True)
        c.fill  = sect
        c.border = thin
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"Specialty Games 告警分析  |  {ana['target_date']}"
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.fill  = navy
    t.border = thin
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    r = 2
    shdr(r, "查询日期当日数据"); r += 1
    wrow(r, "查询日期",  ana["target_date"]); r += 1
    wrow(r, "时间段",    f"{start_time} ~ {end_time}"); r += 1
    wrow(r, "BET",       round(ana["cur_bet"],    3)); r += 1
    wrow(r, "PAYOUT",    round(ana["cur_payout"], 3)); r += 1
    wrow(r, "RTP",       f"{ana['cur_rtp']:.4f}%"); r += 1
    wrow(r, "当日 GGR (totalBonusAmount)", round(ana["cur_ggr"], 3)); r += 1

    r += 1; shdr(r, f"前 {ana['days_back']} 天分析"); r += 1
    wrow(r, f"前{ana['days_back']}天日均盈利率",  f"{ana['avg_nd_rtp']:.4f}%"); r += 1
    wrow(r, f"前{ana['days_back']}天投注额中位数", f"{ana['median_nd_bet']:,.3f}"); r += 1
    wrow(r, "当前RTP − 日均RTP",  f"{ana['rtp_diff']:.4f}%"); r += 1
    wrow(r, "告警阈值 X",         f"{ana['x_threshold']}%"); r += 1

    r += 1; shdr(r, f"前{ana['days_back']}天 Sorted BET"); r += 1
    for idx, b in enumerate(ana["sorted_bets"], 1):
        ws.cell(row=r, column=1, value=f"第{idx}天").border = thin
        c2 = ws.cell(row=r, column=2, value=round(b, 3))
        c2.border = thin
        c2.number_format = "#,##0.000"
        r += 1

    r += 1; shdr(r, "告警判断"); r += 1
    a1f = red if ana["alert1"] else green
    a2f = red if ana["alert2"] else green
    a3f = red if ana["alert3"] else green
    nf  = red if ana["normal_alert"] else green
    cf  = red if ana["consecutive_alert"] else green
    wrow(r, f"条件1: RTP差值({ana['rtp_diff']:.4f}%) < X({ana['x_threshold']}%)",
         "TRUE ⚠" if ana["alert1"] else "FALSE ✓", fill=a1f, bold=True); r += 1
    wrow(r, f"条件2: BET({ana['cur_bet']:,.0f}) ≥ 中位数({ana['median_nd_bet']:,.0f})",
         "TRUE ⚠" if ana["alert2"] else "FALSE ✓", fill=a2f, bold=True); r += 1
    wrow(r, (f"条件3: 上一GGR({ana['prev_ggr']:,.3f}) − 当日GGR({ana['cur_ggr']:,.3f})"
             f" = {ana['c3_diff']:,.3f} ≥ |上一GGR|×Y({ana['y_threshold']}%) = {ana['c3_abs']:,.3f}"),
         "TRUE ⚠" if ana["alert3"] else "FALSE ✓", fill=a3f, bold=True); r += 1
    r += 1
    wrow(r, "普通告警（条件1 AND 条件2）",
         "TRUE ⚠" if ana["normal_alert"] else "FALSE ✓", fill=nf, bold=True); r += 1
    wrow(r, "连续告警（条件1 AND 条件2 AND 条件3）",
         "TRUE ⚠" if ana["consecutive_alert"] else "FALSE ✓", fill=cf, bold=True)


# ─── 主流程 ───────────────────────────────────────────────────────────────────
async def run(query_date: date, start_time: str, end_time: str,
              days_back: int, outlet_ids: list, game_brands: list,
              x_threshold: float, y_threshold: float, prev_ggr: float,
              export_raw: bool, export_analysis: bool,
              username: str, password: str, output: str) -> None:

    all_dates  = [query_date - timedelta(days=i) for i in range(days_back, -1, -1)]
    past_dates = all_dates[:-1]

    plog(f"查询日期: {query_date}  时间段: {start_time} ~ {end_time}")
    plog(f"历史天数: {days_back}  Outlets: {outlet_ids}  Brands: {game_brands}")
    plog(f"X={x_threshold}%  Y={y_threshold}%  prevGGR={prev_ggr}")

    # Phase 1
    plog("\n── Phase 1: 浏览器抓取首日数据 ──")
    p1 = await phase1_capture(query_date, start_time, end_time,
                              outlet_ids, game_brands, username, password)

    # Phase 2
    plog(f"\n── Phase 2: 并发复放历史 {len(past_dates)} 天请求 ──")
    p2_results = await phase2_parallel(p1, past_dates, query_date)

    # 整合所有天数数据
    rows = []
    for day in all_dates:
        ds      = str(day)
        summary = p1.get("first_data") if day == query_date else p2_results.get(ds)
        if summary:
            row = {"日期": ds, "时间段": f"{start_time}~{end_time}", "状态": "✓"}
            for f in NUMERIC_FIELDS:
                row[f] = summary.get(f, 0)
        else:
            row = {"日期": ds, "时间段": f"{start_time}~{end_time}", "状态": "无数据"}
            for f in NUMERIC_FIELDS:
                row[f] = ""
        rows.append(row)

    ok = sum(1 for r in rows if r["状态"] == "✓")
    plog(f"\n共 {len(rows)} 天，成功 {ok} 天，无数据 {len(rows) - ok} 天")

    # 计算告警分析
    ana = compute_analysis(rows, str(query_date), x_threshold, y_threshold, prev_ggr, days_back)

    # 导出 Excel
    wb = openpyxl.Workbook()
    first_sheet = True

    if export_raw:
        if first_sheet:
            ws = wb.active; ws.title = "原始数据"; first_sheet = False
        else:
            ws = wb.create_sheet("原始数据")
        write_raw_sheet(ws, rows, start_time, end_time)
        plog("✅ Sheet「原始数据」已写入")

    if export_analysis:
        if first_sheet:
            ws = wb.active; ws.title = "每日数据"; first_sheet = False
        else:
            ws = wb.create_sheet("每日数据")
        write_daily_sheet(ws, rows, ana, start_time, end_time)

        ws2 = wb.create_sheet("告警分析")
        write_analysis_sheet(ws2, ana, start_time, end_time)
        plog("✅ Sheet「每日数据」+「告警分析」已写入")

    if first_sheet:
        # 没有选任何导出项，默认写每日数据
        ws = wb.active; ws.title = "每日数据"
        write_daily_sheet(ws, rows, ana, start_time, end_time)

    out_dir = os.path.dirname(os.path.abspath(output))
    os.makedirs(out_dir, exist_ok=True)
    try:
        wb.save(output)
    except Exception as e:
        err_exit(f"写入 Excel 失败: {e}")

    plog(f"✨ 导出完成  → {output}")
    print(json.dumps({
        "success":    True,
        "rows":       ok,
        "output":     output,
        "alert1":     ana["alert1"],
        "alert2":     ana["alert2"],
        "alert3":     ana["alert3"],
        "normalAlert":      ana["normal_alert"],
        "consecutiveAlert": ana["consecutive_alert"],
    }, ensure_ascii=False))


# ─── CLI 入口 ─────────────────────────────────────────────────────────────────
def calc_start_time(end_time_str: str) -> str:
    """结束时间往前推 3599 秒（约 1 小时）"""
    t = datetime.strptime(end_time_str, "%H:%M:%S")
    s = (datetime.combine(date.today(), t.time()) - timedelta(seconds=3599)).time()
    return s.strftime("%H:%M:%S")


def main():
    parser = argparse.ArgumentParser(description="IGO Specialty Games 导出工具")
    parser.add_argument("--query-date",      default="",
                        help="查询日期 YYYY-MM-DD（默认今天）")
    parser.add_argument("--end-time",        default="11:00:59",
                        help="结束时间 HH:MM:SS（默认 11:00:59；开始时间自动推算为 end-3599s）")
    parser.add_argument("--days-back",       type=int, default=30,
                        help="回溯天数（默认 30）")
    parser.add_argument("--outlet-ids",      default="IGO,Lavie,NCLI,Stotsenberg",
                        help="Outlet ID，逗号分隔")
    parser.add_argument("--game-brands",     default="IGO-COLORGAME,QAT-55-COLORGAME,QAT-COLORGAME",
                        help="Game Brand，逗号分隔")
    parser.add_argument("--x-threshold",     type=float, default=5.0,
                        help="RTP 差值告警阈值 X%%（默认 5.0）")
    parser.add_argument("--y-threshold",     type=float, default=10.0,
                        help="连续告警 GGR 变化阈值 Y%%（默认 10.0）")
    parser.add_argument("--prev-ggr",        type=float, default=0.0,
                        help="上次普通告警时的 GGR（即 totalBonusAmount 取反，默认 0）")
    parser.add_argument("--export-raw",      action="store_true",
                        help="导出原始数据 sheet")
    parser.add_argument("--export-analysis", action="store_true",
                        help="导出每日数据 + 告警分析 sheet（默认 true 当无其他 flag 时）")
    parser.add_argument("--output",          required=True,
                        help="输出 Excel 路径")
    parser.add_argument("--username",        default=DEFAULT_USERNAME,
                        help="IGO 账号")
    parser.add_argument("--password",        default=DEFAULT_PASSWORD,
                        help="IGO 密码")
    args = parser.parse_args()

    if args.query_date:
        try:
            query_date = datetime.strptime(args.query_date, "%Y-%m-%d").date()
        except ValueError:
            err_exit(f"日期格式错误: {args.query_date}", "请使用 YYYY-MM-DD")
    else:
        query_date = date.today()

    # 默认：若两个导出 flag 均未设置，则自动启用 export_analysis
    export_raw      = args.export_raw
    export_analysis = args.export_analysis
    if not export_raw and not export_analysis:
        export_analysis = True

    try:
        datetime.strptime(args.end_time, "%H:%M:%S")
    except ValueError:
        err_exit(f"结束时间格式错误: {args.end_time}", "请使用 HH:MM:SS")

    start_time = calc_start_time(args.end_time)
    end_time   = args.end_time

    outlet_ids  = [x.strip() for x in args.outlet_ids.split(",")  if x.strip()]
    game_brands = [x.strip() for x in args.game_brands.split(",") if x.strip()]

    asyncio.run(run(
        query_date      = query_date,
        start_time      = start_time,
        end_time        = end_time,
        days_back       = max(1, args.days_back),
        outlet_ids      = outlet_ids,
        game_brands     = game_brands,
        x_threshold     = args.x_threshold,
        y_threshold     = args.y_threshold,
        prev_ggr        = args.prev_ggr,
        export_raw      = export_raw,
        export_analysis = export_analysis,
        username        = args.username,
        password        = args.password,
        output          = args.output,
    ))


if __name__ == "__main__":
    main()
